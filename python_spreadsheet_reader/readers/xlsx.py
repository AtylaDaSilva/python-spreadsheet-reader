from python_spreadsheet_reader.readers.exceptions import SpreadsheetIsLockedException, NoActiveSpreadsheetException
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.drawing.image import Image
from pathlib import Path
from typing import Any
import os
import re


class XLSXReader:
    def __init__(self, workbook_path: str | Path) -> None:
        self.workbook_path: Path = (
            Path(workbook_path)
            if isinstance(workbook_path, str)
            else workbook_path
        )
        self._sheet_data: dict[int, dict] = {}  # Local cache
        self._workbook: openpyxl.Workbook | None = None

    # * ---------------------------------------------------------------------------
    # *                               Public API
    # * ---------------------------------------------------------------------------
    
    def load_workbook(
            self,
            read_only: bool = False,
            keep_vba: bool = False,
            keep_formulae: bool = True,
            keep_links: bool = True,
            keep_rich_text: bool = False,
            read_locked: bool = False
    ) -> openpyxl.Workbook:
        """
        Opens and returns the workbook.
        Args:
            read_only:
                Opens the workbook in an optimized for reading mode, but content can't be edited. Defaults to True.
            keep_vba:
                If True, preserves VBA content (this does NOT mean you can use it). Defaults to False.
            keep_formulae:
                If True, returns cell formulae instead of cached values. Defaults to True.
            keep_links:
                If True, preserves links to external workbooks. Defaults to True.
            keep_rich_text:
                If True, preserves any rich text formatting in cells. Defaults to False.
            read_locked:
                If True, allows the reading of locked (currently opened) spreadsheets. Defaults to False.

        Returns:
            *openpyxl.Workbook* object
        """
        # Validations
        if not self.workbook_path.exists():
            raise FileNotFoundError(
                f'Could not find workbook "{self.workbook_path}"'.replace("\\", "/")
            )
        elif not read_locked and self.is_spreadsheet_locked():
            raise SpreadsheetIsLockedException(self.workbook_path)
        if self._workbook is not None:
            return self._workbook
        wb = openpyxl.load_workbook(
            filename=self.workbook_path,
            read_only=read_only,
            keep_vba=keep_vba,
            data_only=(not keep_formulae),
            keep_links=keep_links,
            rich_text=keep_rich_text
        )
        self._workbook = wb
        return wb

    def close_workbook(self):
        if self._workbook:
            self._workbook.close()
            self._workbook = None
            self._sheet_data = {}

    def read_sheet(
            self, sheet_name: str | None = None,
            cell_values_only: bool = False,
            return_cell_coords: bool = True,
            read_only: bool = False,
            keep_vba: bool = False,
            keep_formulae: bool = True,
            keep_links: bool = True,
            keep_rich_text: bool = False,
            read_locked: bool = False,
            close_workbook: bool = True,
    ) -> dict[int, dict]:
        """
        Returns the data from the spreadsheet located at *self.workbook_path*.
        Args:
            sheet_name:
                Name of the sheet to read data from. If not specified, returns the data from the active sheet.
            cell_values_only:
                If True, returns cell values (str, int, datetime, etc.) instead of Cell objects. Defaults to False.
            return_cell_coords:
                If True, returns cell coordinates (A1, B2, C3, ect.) instead of column numbers.
                Does not affect row numbers. Defaults to True.
            read_only:
                Opens the workbook in an optimized for reading mode, but content can't be edited. Defaults to True.
            keep_vba:
                If True, preserves VBA content (this does NOT mean you can use it). Defaults to False.
            keep_formulae:
                If True, returns cell formulae instead of cached values. Defaults to True.
            keep_links:
                If True, preserves links to external workbooks. Defaults to True.
            keep_rich_text:
                If True, preserves any rich text formatting in cells. Defaults to False.
            read_locked:
                If True, allows the reading of locked (currently opened) spreadsheets. Defaults to False.
            close_workbook:
                If True, closes the workbook after returning the data. Defaults to True.

        Returns: A dict of sheet rows, each key representing the row number (1-based) and each value a nested dict.
        The nested dicts represents cells, with cell coordinates (or column numbers) as keys and cell values (or objects) as values.

        Examples:
            >>>reader = XLSXReader(workbook_path="path/to/workbook.xlsx")
            >>>sheet_data = reader.read_sheet("Sheet2", cell_values_only=True)

            >>>print(sheet_name)

            # Outputs

            {
                1: {"A1", "ID", "A2": "Title", "A3": "Genre"},

                2: {"B1", 123, "B2": "Alien", "A3": "Science Fiction, Horror"},

                2: {"C1", 124, "C2": "Predator", "A3": "Science Fiction, Action"},

                ...
            }

        """
        match self.workbook_path.suffix.lower():
            case ".xlsx":
                # Open workbook
                self.load_workbook(
                    read_only,
                    keep_vba,
                    keep_formulae,
                    keep_links,
                    keep_rich_text,
                    read_locked,
                )
                # Get active (or specific, if provided) spreadsheet
                ws = self._get_worksheet(sheet_name)

                # Iterate through rows/cols to get cell values
                rows = {}
                for i, row in enumerate(ws.iter_rows()):
                    r = {}
                    for cell in row:
                        if isinstance(cell, openpyxl.cell.read_only.EmptyCell):  # Ignore empty cells
                            continue
                        r[cell.coordinate if return_cell_coords else cell.column] = cell.value if cell_values_only else cell
                    if len(r) > 0:  # Ignore empty rows
                        rows[i + 1] = r
                if close_workbook:
                    self.close_workbook()
                    return rows
                self._sheet_data = rows
                return self._sheet_data
            case _:
                raise ValueError(f'Unsupported file type: "{self.workbook_path.suffix}".')

    def get_cell(
            self,
            coords: str | None = None,
            row: int | None = None,
            col: int | None = None,
            sheet_name: str | None = None
    ) -> Cell | MergedCell:
        ws = self._get_worksheet(sheet_name)

        if coords:
            return ws[coords]
        elif row and col:
            return ws.cell(row=row, column=col)
        else:
            raise ValueError(f"Provide either cell coordinates or row/col indices.")

    def set_cell_value(
            self,
            value: Any,
            coords: str | None = None,
            row: int | None = None,
            col: int | None = None,
            sheet_name: str | None = None
    ) -> Cell | MergedCell:
        cell = self.get_cell(coords, row, col, sheet_name)
        cell.value = value
        return cell

    def adicionar_imagem(
        self, caminho_imagem: str, celula: str, altura: int, largura: int
    ):
        raise NotImplementedError
        imagem = Image(caminho_imagem)
        imagem.height = altura
        imagem.width = largura
        ws = self._planilha_ativa()
        ws.add_image(imagem, celula)

    def save_spreadsheet(self, workbook_path: str | Path | None = None, close_workbook: bool = True) -> Path:
        if workbook_path:
            p: Path = Path(workbook_path) if isinstance(workbook_path, str) else workbook_path
        else:
            p: Path = self.workbook_path
        p.parent.mkdir(parents=True, exist_ok=True)
        wb = self.load_workbook()
        wb.save(p)
        if close_workbook:
            self.close_workbook()
        return p

    def is_spreadsheet_locked(self) -> bool:
        """
        Return True if the spreadsheet at filepath is currently locked by
        another process, False otherwise.

        Checks for lock files created by Excel (~$<filename>) and LibreOffice
        (.~lock.<filename>#) in the same directory as the target file.

        Parameters
        ----------
        filepath:
            Path to the target spreadsheet.

        Raises
        ------
        FileNotFoundError
            If filepath does not exist.
        """
        if not self.workbook_path.exists():
            raise FileNotFoundError(f'Could not find spreadsheet: "{self.workbook_path}"')

        return self._is_xlsx_locked() or self._is_ods_locked()

    # * ---------------------------------------------------------------------------
    # *                                Internal API
    # * ---------------------------------------------------------------------------

    def _get_worksheet(self, sheet_name: str | None = None):
        wb = self.load_workbook()
        if sheet_name:
            return wb[sheet_name]
        active_sheet = wb.active
        if active_sheet is None:
            raise NoActiveSpreadsheetException(self.workbook_path)
        return active_sheet

    def _is_xlsx_locked(self) -> bool:
        """
        Return True if an Excel lock file (~$<filename>) exists in the same
        directory as filepath, False otherwise.

        Excel creates a temporary lock file prefixed with "~$" when a workbook
        is open. Its presence indicates the file is currently in use.
        """
        return bool(re.search(r"~\$" + re.escape(self.workbook_path.name), " ".join(
            f.name for f in self.workbook_path.parent.iterdir()
        )))

    def _is_ods_locked(self) -> bool:
        """
        Return True if a LibreOffice lock file (.~lock.<filename>#) exists in
        the same directory as filepath, False otherwise.

        LibreOffice creates a lock file with the pattern .~lock.<filename># when
        a document is open. Its presence indicates the file is currently in use.
        """
        return bool(re.search(r"\.~lock\." + re.escape(self.workbook_path.name) + r"#", " ".join(
            f.name for f in self.workbook_path.parent.iterdir()
        )))
