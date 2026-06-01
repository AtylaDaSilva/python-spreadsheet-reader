import openpyxl
from openpyxl.styles import Alignment
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.drawing.image import Image
from pathlib import Path
from typing import Any
import os
import re

# ===========================================================================================
#                                      Leitor Planilha
# ===========================================================================================


class LeitorXLSX:
    def __init__(self, caminho_planilha: str | Path, *args, **kwargs) -> None:
        self.caminho_planilha = (
            Path(caminho_planilha)
            if isinstance(caminho_planilha, str)
            else caminho_planilha
        )
        self._sheet_data: dict[int, tuple] = {}
        self._workbook: openpyxl.Workbook | None = None
        self._abrir_workbook(*args, **kwargs)

    # ---------------------------------------------------------------------------
    # API Pública
    # ---------------------------------------------------------------------------

    def ler_excel(self) -> dict[int, tuple]:
        """
        Read an Excel spreadsheet and return its rows as a dict of tuples.

        Returns:
            A dict of tuples, each key representing the column number (not 0-indexed).

        Raises:
            FileNotFoundError: If the file does not exist.
            ValueError:        If the file extension is not .xlsx.
        """
        file = self.caminho_planilha

        if not file.exists():
            msg = f"Arquivo não encontrado: {file}"
            raise FileNotFoundError(
                msg.replace("\\", "/")  # Formatar \\ no caminho do arquivo
            )

        match file.suffix.lower():
            case ".xlsx":
                return self._read_xlsx()
            case _:
                raise ValueError(f"Tipo de arquivo '{file.suffix}' não suportado.")

    def get_celula(self, linha: int, coluna: int) -> Cell | MergedCell:
        ws = self._planilha_ativa()
        return ws.cell(row=linha, column=coluna)

    def get_valor_celula(self, linha: int, coluna: int) -> Any:
        return self.get_celula(linha=linha, coluna=coluna).value

    def set_valor_celula(self, linha: int, coluna: int, valor: Any) -> None:
        self.get_celula(linha=linha, coluna=coluna).value = valor

    def adicionar_imagem(
        self, caminho_imagem: str, celula: str, altura: int, largura: int
    ):
        imagem = Image(caminho_imagem)
        imagem.height = altura
        imagem.width = largura
        ws = self._planilha_ativa()
        ws.add_image(imagem, celula)

    def salvar_planilha(self, caminho: str | Path | None = None) -> Path:
        wb = self._abrir_workbook()
        p: Path = self.caminho_planilha if not caminho else Path(caminho)
        if not p.parent.is_dir():
            p.parent.mkdir(parents=True, exist_ok=True)
        wb.save(p)
        return p

    def fechar_planilha(self):
        self._fechar_workbook()

    # ---------------------------------------------------------------------------
    # Métodos Protegidos
    # ---------------------------------------------------------------------------

    def _abrir_workbook(self, *args, **kwargs) -> openpyxl.Workbook:
        if self.is_spreadsheet_locked():
            raise SpreadsheetIsLockedException(self.caminho_planilha)
        if self._workbook is not None:
            return self._workbook
        wb = openpyxl.load_workbook(self.caminho_planilha, *args, **kwargs)
        self._workbook = wb
        return wb

    def _fechar_workbook(self):
        if self._workbook:
            self._workbook.close()
            self._workbook = None
            self._sheet_data = {}

    def _planilha_ativa(self):
        active_sheet = self._abrir_workbook().active
        if active_sheet is None:
            raise NoActiveSheetException(
                f"Não foi possível localizar a planilha ativa no excel {self.caminho_planilha}"
            )
        return active_sheet

    def _read_xlsx(self) -> dict[int, tuple]:
        ws = self._planilha_ativa()
        rows = {}
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            rows[i + 1] = row
        self._sheet_data = rows
        return self._sheet_data

    def _is_xlsx_locked(self) -> bool:
        """
        Return True if an Excel lock file (~$<filename>) exists in the same
        directory as filepath, False otherwise.

        Excel creates a temporary lock file prefixed with "~$" when a workbook
        is open. Its presence indicates the file is currently in use.
        """
        return bool(re.search(r"~\$" + re.escape(self.caminho_planilha.name), " ".join(
            f.name for f in self.caminho_planilha.parent.iterdir()
        )))

    def _is_ods_locked(self) -> bool:
        """
        Return True if a LibreOffice lock file (.~lock.<filename>#) exists in
        the same directory as filepath, False otherwise.

        LibreOffice creates a lock file with the pattern .~lock.<filename># when
        a document is open. Its presence indicates the file is currently in use.
        """
        return bool(re.search(r"\.~lock\." + re.escape(self.caminho_planilha.name) + r"#", " ".join(
            f.name for f in self.caminho_planilha.parent.iterdir()
        )))

    def is_spreadsheet_locked(self) -> bool:
        """
        Return True if the spreadsheet at filepath is currently locked by
        another process, False otherwise.

        Checks for lock files created by Excel (~$<filename>) and LibreOffice
        (.~lock.<filename>#) in the same directory as the target file.

        Parameters
        ----------
        filepath:
            Path to the target spreadsheet.

        Raises
        ------
        FileNotFoundError
            If filepath does not exist.
        """

        if not self.caminho_planilha.exists():
            raise FileNotFoundError(f"Não foi possível encontrar a planilha: {self.caminho_planilha}")

        return self._is_xlsx_locked() or self._is_ods_locked()

# ===========================================================================================
#                                        Helpers / Outras classes
# ===========================================================================================


class EstilizadorDeCelula:
    def __init__(self, celula: Cell | MergedCell) -> None:
        self.celula = celula

    def alinhar(self, horizontal: str, vertical: str, quebrar_texo: bool):
        self.celula.alignment = Alignment(
            horizontal=horizontal, vertical=vertical, wrapText=quebrar_texo
        )

    def formatar(self, formato: str) -> None:
        self.celula.number_format = formato


# ===========================================================================================
#                                        Exceptions
# ===========================================================================================


class NoActiveSheetException(Exception):
    pass


class SpreadsheetIsLockedException(Exception):
    pass
